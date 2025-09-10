from django.shortcuts import render ,  redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import vc_n_asn, VcMaster, VcDatabase, EslPart, AsnSchedule, WorkTable, trolley_data, buffer
import requests
from django.contrib import messages
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Count, Exists, OuterRef, Q
import json
from itertools import cycle
from django.template.loader import render_to_string
from django.http import HttpResponseBadRequest
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from collections import defaultdict
import openpyxl
from django.core.files.storage import FileSystemStorage
from django.views.decorators.http import require_http_methods
from django.db import transaction
import os
from datetime import datetime


# Initialize an in-memory queue
pending_data_queue = defaultdict(list)

url1 = 'http://192.168.1.200/wms/associate/updateScreen'
styleMap_dict = {
     "esl_styleid":51,
    "esl_maptype":591,
    "trolley_styleid":50,
    "trolley_maptype":412,
}

# Assuming your trolley model is `Trolley`

# Define a list of queue colors
seven_queue = ["ff00", "ffff00", "ff", "ff0000", "ff00ff", "ffffff", "00ffff"]

# Function to display the color code
def display_color():
    # Get the first color code from the list
    color_code = seven_queue.pop(0)

    # Add the color code back to the end of the list
    seven_queue.append(color_code)

    # Return the color code
    return color_code


global vc_number
posted_vc = None

from django.contrib.messages import get_messages


@csrf_exempt
def get_messages_view(request):
    storage = get_messages(request)
    messages = [{'level': message.level, 'message': message.message} for message in storage]
    return JsonResponse({'messages': messages})

@csrf_exempt
def get_combined_data(request):
    qr_data = request.POST.get('qr_data')
    vc_n_asn_data = vc_n_asn.objects.all()
    combined_data = []
    print('combined Data:', qr_data)
    for entry in vc_n_asn_data:
        vc_number = entry.vcn
        asn_number = entry.asnn
        matching_models = VcMaster.objects.filter(vcnumber=vc_number)

        if matching_models.exists():
            model_info = matching_models.first().model
        else:
            model_info = 'No matching model found'

        data_entry = {
            'vc_number': vc_number,
            'asn_number': asn_number,
            'model': model_info,
            'trolley_qr': qr_data,
        }
        combined_data.append(data_entry)
        kitting_in_process_data = [entry for entry in combined_data if entry['vc_number'] == posted_vc]
        print('psted_vc', posted_vc)
    return JsonResponse({'combined_data': combined_data, 'kitting_in_process_data': kitting_in_process_data})


@csrf_exempt
def picking_plan(request):
    global posted_vc
    # vc_n_asn.objects.filter(vcn = posted_vc).delete()
    vc_n_asn_data = vc_n_asn.objects.order_by('-schedule_date_time').filter(status=0)

    # Define a list to store the combined data

    combined_data = []

    # Iterate over each entry in vc_n_asn_data
    for entry in vc_n_asn_data:
        vc_number = entry.vcn
        asn_number = entry.asnn
        date_time = entry.schedule_date_time

        # Query VcMaster to find matching model for the VC number
        matching_models = VcMaster.objects.filter(vcnumber=vc_number)

        # Check if matching_models is not empty
        if matching_models.exists():
            # Retrieve the first matching model
            model_info = matching_models.first().model

            # Create a dictionary to store VC, ASN, and model information
            data_entry = {
                'vc_number': vc_number,
                'asn_number': asn_number,
                'model': model_info,
                'plan_date': date_time.date(),
                # 'schedule_time': datetime.strftime(date_time, '%-I:%-M %p'),
                'schedule_time': datetime.strftime(date_time, '%c'),
            }

            # Add the dictionary to the combined_data list
            combined_data.append(data_entry)
            # combined_data = [entry for entry in combined_data if entry['vc_number'] != posted_vc]
            combined_data = [entry for entry in combined_data]

            # add code to delete the combined data entry if match posted vc
        else:
            # If no matching model found, append a message to the combined_data list
            combined_data.append({
                'vc_number': vc_number,
                'asn_number': asn_number,
                'model': 'No matching model found',
                'plan_date': date_time.date(),
                'schedule_time': date_time.time(),
            })

             # Create a Paginator object with 10 items per page
    paginator = Paginator(combined_data, 15)

    # Get the current page number from the request, defaulting to 1
    page_number = request.GET.get('page', 1)

    try:
        # Get the requested page
        page = paginator.page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        page = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results
        page = paginator.page(paginator.num_pages)

    # Pass the paginated data to the template
    return render(request, 'pick_plan.html', {'page': page})

    # return render(request, 'pick_plan.html', {'combined_data': combined_data})

#In this we create payload and append to data list , in getpayload we will post this payload to esl
def prepare_data_list(vc_number , vc_data_list , qr_data): 
    #data_list = []
    # Create a dictionary to hold multiple data lists based on router numbers
    data_lists = defaultdict(list)
    vc_color_mapping = {}
    vc_color = vc_color_mapping.get(vc_number)
    if not vc_color:
        vc_color = display_color()
        vc_color_mapping[vc_number] = vc_color

    for vc_data in vc_data_list:
        part_number = vc_data.part_no
        try:
            esl_data = EslPart.objects.get(partno=part_number)
            # Dynamically determine the router number
            router_number= esl_data.router_numb
            data_lists[router_number].append({
                "mac": esl_data.tagid,
                "mappingtype": styleMap_dict['esl_maptype'],
                "Part No.": part_number,
                "DESC": vc_data.part_desc,
                "QTY": vc_data.quantity,
                "ledstate": "0",
                "ledrgb": vc_color,
                "outtime": "0",
                "styleid": styleMap_dict['esl_styleid'],
                "qrcode": qr_data,# change this with matching trolley
            })
        except EslPart.DoesNotExist:
            pass
    # Print lengths of the generated data lists for debugging
    for router_number, data_list in data_lists.items():
        print(f"Data list for Router {router_number}: {len(data_list)} entries")

    return data_lists, vc_color 

def post_data_to_url(url, data_list, vc_number, qr_data, vc_color):
    """
    Posts data to the specified URL and handles the response.
    
    Args:
        url (str): The URL to post data to.
        data_list (list): The list of data to be posted.
        vc_number (str): The VC number associated with the data.
        qr_data (str): QR code data.
        vc_color (str): The color mapping for the VC.

    Returns:
        dict: Contains 'success' or 'error' key with a message.
    """
    try:
        # Post the data to the specified URL
        response = requests.post(url, json=data_list)
        response.raise_for_status()  # Raise an error if the response is not successful

        if not response.ok:
            return {'error': f'Failed to post data to URL {url}'}

        # Retrieve ASN number and model info
        asn_number, model_info = get_vc_details(vc_number)
        if not asn_number or not model_info:
            return {'error': 'No ASN or model description found'}

        # Create ASN schedule
        asn_schedule_created = create_asn_schedule(vc_number, asn_number, model_info, qr_data, vc_color)
        if not asn_schedule_created:
            return {'error': 'ASN Schedule not created'}

        # Update trolley status
        update_trolley_status(qr_data, asn_number, vc_color, model_info)

        # Create entries in the WorkTable
        create_worktable_entries(data_list, asn_number, vc_color)

        return {'success': f'Data successfully posted to {url} and processed.'}
    
    except Exception as e:
        return {'error': str(e)}
   

def post_data_to_server(url, list_of_data, vc_number, qr_data, vc_color):
    """
    Posts data to the specified URL and handles the response.
    
    Args:
        url (str): The URL to post data to.
        data_list (list): The list of data to be posted.
        vc_number (str): The VC number associated with the data.
        qr_data (str): QR code data.
        vc_color (str): The color mapping for the VC.

    Returns:
        dict: Contains 'success' or 'error' key with a message.
    """
    # if list_of_data:
    #     for list_item in list_of_data:
    #         router_number = list_item.router_numb
    try:
        # Post the data to the specified URL
        response = requests.post(url, json=data_list)
        response.raise_for_status()  # Raise an error if the response is not successful

        if not response.ok:
            return {'error': f'Failed to post data to URL {url}'}

        # Retrieve ASN number and model info
        asn_number, model_info = get_vc_details(vc_number)
        if not asn_number or not model_info:
            return {'error': 'No ASN or model description found'}

        # Create ASN schedule
        asn_schedule_created = create_asn_schedule(vc_number, asn_number, model_info, qr_data, vc_color)
        if not asn_schedule_created:
            return {'error': 'ASN Schedule not created'}

        

        return {'success': f'Data successfully posted to {url} and processed.'}
    
    except Exception as e:
        return {'error': str(e)}
    

def delete_repeat_asn_data_helper(asn_number):
    # This is a regular Python function, not a Django view
    # It handles the deletion logic and returns the results
    if not asn_number:
        return {'status': 'error', 'message': 'ASN number not provided'}

    try:
        
        repeated_asn = AsnSchedule.objects.filter(asn_no=asn_number)
        if repeated_asn:

            deleted_asnschedules =  AsnSchedule.objects.all().delete()
            deleted_worktables  = WorkTable.objects.all().delete()
        # Add logic to delete entries from vc_n_asn except the latest 999
        # 1. Order by date/time to find the latest entries
        latest_entries = vc_n_asn.objects.order_by('-schedule_date_time')[:999]
        # 2. Get the IDs of these latest entries
        latest_ids = [entry.id for entry in latest_entries]
        # 3. Delete all entries whose IDs are NOT in the latest_ids list
        vc_n_asn.objects.exclude(id__in=latest_ids).delete()

        
        return {
            'status': 'success',
            'worktable_deleted': deleted_worktables,
            'asnschedule_deleted': deleted_asnschedules,
            'message': f'Data for ASN {asn_number} deleted successfully.'
        }
    except Exception as e:
        # Proper error handling for the deletion process
        return {'status': 'error', 'message': f'Error deleting ASN data: {e}'}


@csrf_exempt  # Disable CSRF protection for this view to allow unauthenticated access
@transaction.atomic  # Make the entire function run in a single database transaction
def get_Payload_Data(request):
    """
   Functionality:
    This function handles a POST request to process QR data or VC number (which represents some logistical information).
    It checks for the availability of trolleys, validates data from the request, posts data to an external URL, and
    updates the status of trolleys and work items. It also maintains a pending data queue for items that haven't been
    processed yet and manages the creation of entries in several tables related to the workflow.

    The function ensures that:
    - Trolleys are available for use and not already engaged.
    - Posts data to an external service if valid.
    - Updates the WorkTable with the data processed.
    - Handles pending and completed items separately, processing each accordingly.
    """

    # Global variables used in this function
    global posted_vc
    global vc_number
    global dt

    # Process only POST requests
    if request.method == 'POST':
        # Retrieve QR data from the POST request
        qr_data = request.POST.get('qr_data')
        print(qr_data)  # Log QR data for debugging purposes


        if qr_data is None:
            vc_number = request.POST.get('vc_number')
            for_repeat_asn = request.POST.get('asn_number')
            
            d = request.POST.get('date')
            t = request.POST.get('time')
            print(d+t)
            dt = datetime.strptime(t, "%c")

            print('VCNUMBER', vc_number, 'dt', dt, 'asn', for_repeat_asn)
            # Call the helper function and get its response
            delete_result = delete_repeat_asn_data_helper(for_repeat_asn)
            if delete_result['status'] == 'error':
                return JsonResponse(delete_result, status=500)



        try:
            # Check if there are at least 7 trolleys already engaged
            if trolley_data.objects.filter(trolley_picking_status="pending").count() >= 7:
                messages.warning(request, 'This trolley is already engaged!')
                return JsonResponse({'error': 'All trolleys are currently engaged'}, status=400)

            # Retrieve trolley data based on the provided QR code
            matching_trolley = trolley_data.objects.filter(trolley_code=qr_data).first()

            # Check if the trolley is already engaged
            if matching_trolley and matching_trolley.trolley_picking_status == "pending":
                messages.warning(request, 'This trolley is already engaged!')
                return JsonResponse({'error': 'This trolley is already engaged'}, status=400)

            # Check if the trolley is invalid
            if not matching_trolley:
                messages.error(request, 'Invalid QR data or Trolley code')
                return JsonResponse({'error': 'Invalid QR data or Trolley code'}, status=404)

            # Check if the trolley status is not completed
            if matching_trolley and matching_trolley.trolley_picking_status != 'completed':
                messages.error(request, 'Invalid QR data or Trolley code')
                return JsonResponse({'error': 'Invalid QR data or Trolley code'}, status=404)

            # Retrieve the list of VC data entries based on the provided VC number
            vc_data_list = VcDatabase.objects.filter(vc_no=vc_number)
            data_lists, vc_color = prepare_data_list(vc_number, vc_data_list, qr_data)  # Prepare data list for processing

            print('datalist', data_list, qr_data, vc_number)  # Log data list, QR data, and VC number for debugging

            # Check if data list and QR data are available
            if data_lists and qr_data:
                for router_number, data_list in data_lists.items():
                    # Skip if the data list is empty
                    if not data_list:
                        continue

                    # Define the URL dynamically based on the router number
                    url_map = {
                        1: "https://example.com/router1",
                        2: "https://example.com/router2",
                        3: "https://example.com/router3",
                    }
                    url = url_map.get(router_number)
                    if not url:
                        continue  # Skip if no URL is defined for this router number

                    # Call the separate function to post data to the URL
                    result = post_data_to_url(url, data_list, vc_number, qr_data, vc_color)

                    # Check the result and return appropriate JSON responses or messages
                    if 'error' in result:
                        messages.error(request, result['error'])
                        return JsonResponse({'error': result['error']}, status=500)
                    else:
                        messages.success(request, result['success'])


                else:
                    pending_items = []  # List to store pending items
                    completed_items = []  # List to store completed items
                    processed_macs = set()  # Set to track processed MAC addresses

                    # Loop through each item in the data list
                    for item in data_list:
                        mac = item['mac']

                        # Retrieve worktable entries for the current MAC address
                        worktable_entries = WorkTable.objects.filter(tagid=mac).order_by('-stdatetime')
                        if worktable_entries:
                            for entry in worktable_entries:
                                if entry.tagid in processed_macs:
                                    continue
                                eslid = str(entry.tagid)

                                # Categorize the worktable entries based on status
                                if entry.status == 'pending':
                                    pending_items.append(item)
                                    pending_data_queue[item['mac']].append(item)
                                    print("pending_data_queue", pending_data_queue)
                                    processed_macs.add(entry.tagid)
                                    posted_vc = vc_number
                                    try:
                                        vc_n_asn_entry = vc_n_asn.objects.get(schedule_date_time=dt)
                                        asn_number = vc_n_asn_entry.asnn
                                        WorkTable.objects.create(
                                            tagid=item['mac'],
                                            tagcode=item['qrcode'],
                                            tagname=item['mac'],
                                            stdatetime=timezone.now(),
                                            partno=item['Part No.'],
                                            partdesc=item['DESC'],
                                            qty=item['QTY'],
                                            asn=asn_number,
                                            color=vc_color,
                                        )
                                    except vc_n_asn.DoesNotExist:
                                        return JsonResponse({'error': 'VC number not found in vc_n_asn'}, status=404)

                                elif entry.status == 'completed':
                                    completed_items.append(item)
                                    processed_macs.add(entry.tagid)

                        else:
                            # If no worktable entry is found, consider it as completed
                            completed_items.append(item)

                    # Check if all items in data list are pending
                    if len(pending_items) == len(data_list):
                        for tagid, items in pending_data_queue.items():
                            asn_number, model_info = get_vc_details(posted_vc)
                            if not asn_number and not model_info:
                                return JsonResponse({'error': 'No ASN or model description found'}, status=404)
                            if WorkTable.objects.filter(asn=asn_number, status='pending'):
                                asn_schedule_created = create_asn_schedule(posted_vc, asn_number, model_info, qr_data, vc_color)
                                if asn_schedule_created:
                                    update_trolley_status(qr_data, asn_schedule_created.asn_no, vc_color, asn_schedule_created.model)
                                    return JsonResponse({'error': 'If all items lie in pending_data_queue, trolley posted successfully'})

                    # Post the completed items if any exist
                    if completed_items:
                        response = requests.post(url1, json=completed_items)
                        response.raise_for_status()
                        posted_vc = vc_number
                        asn_number, model_info = get_vc_details(posted_vc)
                        if not asn_number and not model_info:
                            return JsonResponse({'error': 'No ASN or model description found'}, status=404)

                        asn_schedule_created = create_asn_schedule(posted_vc, asn_number, model_info, qr_data, vc_color)
                        if not asn_schedule_created:
                            return JsonResponse({'error': 'ASN Schedule not created'}, status=500)

                        update_trolley_status(qr_data, asn_schedule_created.asn_no, vc_color, asn_schedule_created.model)
                        create_worktable_entries(completed_items, asn_number, vc_color)

                        return JsonResponse({'success': 'Data posted successfully'})
                    else:
                        messages.warning(request, 'Failed to post data. Check connection with trolley, already engaged!')
                        return JsonResponse({'error': 'Failed to post data'}, status=500)

            else:
                messages.error(request, 'No matching part numbers found in ESL model')
                return JsonResponse({'error': 'No matching part numbers found in ESL model'}, status=404)

        except VcDatabase.DoesNotExist:
            return JsonResponse({'error': 'VC number not found'}, status=404)

        # Fallback response if no conditions are met
        return JsonResponse({'error': 'Unhandled request condition'}, status=500)

    else:
        return JsonResponse({'error': 'Invalid request'}, status=400)

    

# To update trolley status to in process in front end 
def update_trolley_status(qr_data, asn_number, vc_color, model_info):
    try:
        vc_n_asn.objects.filter(schedule_date_time=dt).update(status=1)
        matching_trolley = trolley_data.objects.filter(trolley_code=qr_data).first()
        if matching_trolley:
            matching_trolley.trolley_picking_status = "pending"

            trolley_mac = matching_trolley.mac
            matching_trolley.asn_num = asn_number
            matching_trolley.color = vc_color
            matching_trolley.save()
            trolley_payload = [{
                "mac": trolley_mac, "mappingtype":styleMap_dict['trolley_maptype'], "styleid": styleMap_dict['trolley_styleid'], "qrcode": qr_data,
                "STATUS": "PENDING", "MODEL": model_info,
                "VC": vc_number, "ASN": asn_number,
                "ledrgb": vc_color, "ledstate": "0", "outtime": "0"}]

            response = requests.post(url1, json=trolley_payload)
            response.raise_for_status()
            
        else:
            messages.error(requests , 'No matching trolley found')
            return JsonResponse({'error': 'No matching trolley found'}, status=404)
    except:
        messages.error(requests , 'Could not update ASN status in database')
        return JsonResponse({'error': 'Could not update ASN status in database'}, status=404)
    

def get_vc_details(posted_vc):
     try:
          vc_n_asn_entry = vc_n_asn.objects.filter(schedule_date_time=dt).first()
          vc_master_entry = VcMaster.objects.filter(vcnumber=vc_number).first()
          return vc_n_asn_entry.asnn, vc_master_entry.model
     except(vc_n_asn.DoesNotExist, VcMaster.DoesNotExist):
          return None , None
     
    
def create_asn_schedule(vc_number, asn_number, model_info, qr_data, vc_color ):
    return AsnSchedule.objects.create(
         vc_no=vc_number, asn_no=asn_number, model= model_info,
         start_time=timezone.now(), trqr= qr_data, color=vc_color
    )        

def create_worktable_entries(data_list, asn_number, vc_color):
    for item in data_list:
        WorkTable.objects.create(
            tagid=item['mac'],
            tagcode=item['qrcode'],
            tagname=item['mac'],
            stdatetime=timezone.now(),
            partno=item['Part No.'],
            partdesc=item['DESC'],
            qty=item['QTY'],
            asn=asn_number,
            color= vc_color,
        )


    

def kitting_in_process(request):
    try:
        # Subquery to check if there are any entries with pending status for the same ASN number
        pending_subquery = WorkTable.objects.filter(
            asn=OuterRef('asn'), status='pending'
        )

        # Annotate each ASN number with a flag indicating if any entry has a pending status
        completed_work_table = WorkTable.objects.annotate(
            any_pending=Exists(pending_subquery)
        )

        # Filter to get only those ASN numbers where all entries are completed and none has a pending status
        filtered_work_table = completed_work_table.values('asn').annotate(
            completed_count=Count('id', filter=Q(status='completed')),
        ).filter(
            completed_count=Count('id'),
            any_pending=False,
        )

               # Get the distinct ASN numbers from the filtered queryset
        completed_asn_values = [item['asn'] for item in filtered_work_table]

        # Show ASN schedules where all worktables are completed but selection_status is NOT completed
        kitting_in_process_data = AsnSchedule.objects.filter(
            asn_no__in=completed_asn_values
        ).exclude(selection_status='completed').distinct('asn_no')

        return render(request, 'kitting_in_process.html', {'kitting_in_process_data': kitting_in_process_data})

    except AsnSchedule.DoesNotExist:
        return render(request, 'kitting_in_process.html', {'kitting_in_process_data': None})


@csrf_exempt
def open_modal(request):
    if request.method == 'POST':
        clicked_asn = request.POST.get('asn_number')

        # Retrieve the relevant data from the database
        open_modal_data = list(WorkTable.objects.filter(asn=clicked_asn).values('partno', 'partdesc', 'qty', 'status','asn'))

        # Return the data as JSON response
        return JsonResponse(open_modal_data, safe=False)

    # Return a JsonResponse even for GET requests with an empty list
    return JsonResponse([], safe=False)


@csrf_exempt
def render_modal(request):
    if request.method == 'POST':
        data = json.loads(request.POST.get('open_modal_data', '[]'))
        rendered_html = render_to_string('modal.html', {'open_modal_data': data})
        return HttpResponse(rendered_html)
    else:
        return HttpResponseBadRequest('Invalid request method')


@require_POST  # Ensure this view only handles POST requests
@csrf_exempt  # Disable CSRF protection for this view
@transaction.atomic  # Ensure that database operations are atomic (all or nothing)
def enter_key(request):
    # Parse the incoming JSON data from the request body
    data = json.loads(request.body)
    print(data)  # Log the received data for debugging
    # mac = data['mac']  # Extract the MAC address from the data
    # new_mac = buffer(mac=mac)  # Create a new buffer instance with the MAC address
    # new_mac.save()  # Save the new buffer instance to the database
    return JsonResponse({'message': 'MAC address saved and task dispatched'})

    
#to update esl worktable status on behalf of clicked button mac stored in buffer table
def esl_update(request):
    try:
        # Fetch all buffer entries ordered by their ID
        data = buffer.objects.order_by('id').all()

        for mac_1 in data:
            mac_address = mac_1.mac  # Get the MAC address for the current buffer entry
            
            # Update WorkTable entry based on the MAC address
            worktable_entry = WorkTable.objects.filter(tagid=mac_address, status='pending').order_by('stdatetime').first()
            if worktable_entry:
                worktable_entry.status = 'completed'  # Mark the entry as completed
                worktable_entry.eddatetime = timezone.now()  # Set the end datetime
                worktable_entry.save()  # Save changes to the database
                print(f'Updated WorkTable entry for MAC address: {mac_address}')

            # Handle pending data associated with the MAC address
            if mac_address in pending_data_queue:
                if pending_data_queue[mac_address]:
                    # Prepare the payload for the external service
                    esl_payload = [pending_data_queue[mac_address].pop(0)]
                    print('esl_payload', esl_payload)
                    response = requests.post(url1, json=esl_payload)  # Send POST request with payload
                    print('pending_data_queue', pending_data_queue)
                    if response.status_code == 200:
                        print(f'Data posted successfully for MAC address: {mac_address}')
                    else:
                        print(f'Failed to post data for MAC address: {mac_address}, Status Code: {response.status_code}, Response: {response.text}')
                else:
                    print(f'No data in pending_data_queue for MAC address: {mac_address}')
            else:
                print(f'No data found for MAC address: {mac_address}')

            # Retrieve the latest completed WorkTable entry based on eddatetime
            latest_completed_worktable = WorkTable.objects.filter(tagid=mac_address, status='completed').order_by('-eddatetime').first()
            print(latest_completed_worktable)  # Log the latest completed entry
            
            if latest_completed_worktable:
                asn = latest_completed_worktable.asn  # Get the ASN from the latest completed entry

                # Check all WorkTable entries related to the ASN
                related_worktables = WorkTable.objects.filter(asn=asn)
                all_completed = all(entry.status == 'completed' for entry in related_worktables)  # Check if all related entries are completed

                if all_completed:
                    # Update the AsnSchedule status to 'completed'
                    asn_schedule = AsnSchedule.objects.filter(asn_no=asn).first()
                    if asn_schedule:
                        asn_schedule.selection_status = 'completed'  # Mark ASN schedule as completed
                        asn_schedule.end_time = timezone.now()  # Set end time
                        asn_schedule.save()  # Save changes
                        print(f'ASN Schedule {asn} updated to completed')

                    # Update the trolley status to 'completed' if it matches the ASN
                    matching_trolleys = trolley_data.objects.filter(asn_num=asn)
                    for trolley in matching_trolleys:
                        trolley.trolley_picking_status = 'completed'  # Mark trolley as completed
                        trolley.save()  # Save changes
                        print(f'Trolley {trolley.mac} updated to completed')

                        # Prepare the payload for the trolley
                        trolley_payload = [{
                            "mac": trolley.mac,
                            "STATUS": trolley.trolley_picking_status,
                            "mappingtype": styleMap_dict['trolley_maptype'],
                            "styleid": styleMap_dict['trolley_styleid'],
                            "qrcode": asn_schedule.trqr,
                            "MODEL": asn_schedule.model,
                            "VC": asn_schedule.vc_no,
                            "ASN": asn_schedule.asn_no + ".",
                            "ledrgb": trolley.color,
                            "ledstate": "1",
                            "outtime": "3"
                        }]
                        if trolley.trolley_picking_status == 'completed':
                            # Send the trolley payload to the external service
                            response = requests.post(url1, json=trolley_payload)
                            response.raise_for_status()  # Raise an error for HTTP errors
                            print(f'Trolley data updated successfully, Response: {response.text}', trolley_payload)
                            trolley.asn_num = None  # Clear ASN number for the trolley
                            trolley.save()  # Save changes to the trolley

        # Delete buffer entries with IDs less than or equal to the last processed ID

    except (WorkTable.DoesNotExist, AsnSchedule.DoesNotExist, trolley_data.DoesNotExist) as e:
        print(f'Exception occurred: {e}', "hello me")
        return JsonResponse({'message': 'Request processed for trolley successfully'})
    except requests.exceptions.RequestException as e:
        print(f'Request to external service failed: {e}', "hellotwo")
        return JsonResponse({'message': 'Failed to update trolley data', 'error': str(e)}, status=500)

    return JsonResponse({'message': 'Request processed successfully'})


# to display completed kittings in frontend
def completed_kittings(request):
    try:
        # Subquery to check if there are any entries with pending status for the same ASN number
        pending_subquery = WorkTable.objects.filter(
            asn=OuterRef('asn'), status='pending'
        )

        # Annotate each ASN number with a flag indicating if any entry has a pending status
        completed_work_table = WorkTable.objects.annotate(
            any_pending=Exists(pending_subquery)
        )

        # Filter to get only those ASN numbers where all entries are completed and none has a pending status
        filtered_work_table = completed_work_table.values('asn').annotate(
            completed_count=Count('id', filter=Q(status='completed')),
        ).filter(
            completed_count=Count('id'),
            any_pending=False,
        )

        # Get the distinct ASN numbers from the filtered queryset
        completed_asn_values = [item['asn'] for item in filtered_work_table]

        # Filter AsnSchedule objects based on the completed ASN numbers
        completed_picks = AsnSchedule.objects.filter(asn_no__in=completed_asn_values,
                                                     selection_status='completed').distinct('asn_no')

        return render(request, 'completed_kittings.html', {'completed_picks': completed_picks})

    except AsnSchedule.DoesNotExist:
        # Handle the case where no AsnSchedule objects are found
        return render(request, 'completed_kittings.html', {'completed_picks': None})



# this function is for completing the kitting in process ,manually
@require_POST
@csrf_exempt
def kitting_config(request):
    try:
        data = json.loads(request.body)
        asn_number = data.get('asn_number')
       

        # Update ASN schedule
        asn_schedule = AsnSchedule.objects.filter(asn_no=asn_number).first()
        if asn_schedule:
            asn_schedule.selection_status = 'completed'
            asn_schedule.end_time = timezone.now()
            asn_schedule.save()

            # Update trolley data
            matching_trolleys = trolley_data.objects.filter(asn_num=asn_number)
            for trolley in matching_trolleys:
                trolley.trolley_picking_status = 'completed'
                trolley.save()

                # Form the payload and post it
                trolley_payload = [{
                    "mac": trolley.mac,
                    "mappingtype": styleMap_dict['trolley_maptype'],
                    "styleid": styleMap_dict['trolley_styleid'],
                    "qrcode": asn_schedule.trqr,
                    "STATUS": "Completed",
                    "MODEL": asn_schedule.model,
                    "VC": asn_schedule.vc_no,
                    "ASN": asn_schedule.asn_no + ".",
                    "ledrgb": trolley.color,
                    "ledstate": "1",
                    "outtime": "3"
                }]
                if trolley.trolley_picking_status == 'completed':
                    try:
                        response = requests.post(url1, json=trolley_payload)
                        response.raise_for_status()
                        trolley.asn_num = None
                        trolley.save()
                    except requests.exceptions.RequestException as e:
                        
                        return JsonResponse({'message': 'Failed to update trolley data', 'error': str(e)}, status=500)

        # Fetch and update WorkTable entries
        worktable_entries = WorkTable.objects.filter(asn=asn_number)
        worktable_entries.update(status='completed', eddatetime=timezone.now())
       
        
        if worktable_entries:
            # Send payload to each ESL whose WorkTable entry comes under the given ASN
            for worktable_entry in worktable_entries:
                
                esl_payload = [{
                    "mac": worktable_entry.tagid,
                    "mappingtype": styleMap_dict['esl_maptype'],
                    "styleid": styleMap_dict['esl_styleid'],
                    "qrcode": worktable_entry.tagcode,
                    "Part No.": worktable_entry.partno,
                    "DESC": worktable_entry.partdesc+".",
                    "QTY": worktable_entry.qty,
                    "ledrgb": worktable_entry.color,
                    "ledstate": "1",
                    "outtime": "1"
                },]
                
                
                if worktable_entry.status== 'completed':
                    try:
                        response = requests.post(url1, json=esl_payload)
                        response.raise_for_status()
                       
                    except requests.exceptions.RequestException as e:
                        
                        return JsonResponse({'message': 'Failed to update ESL data', 'error': str(e)}, status=500)

            return JsonResponse({'message': f'ASN:{asn_number} status and related data updated successfully'})
    except (WorkTable.DoesNotExist, AsnSchedule.DoesNotExist, trolley_data.DoesNotExist) as e:
       
        return JsonResponse({'message': 'Error processing ASN', 'error': str(e)}, status=500)
    except json.JSONDecodeError as e:
        
        return JsonResponse({'message': 'Invalid JSON received', 'error': str(e)}, status=400)
    except Exception as e:
       
        return JsonResponse({'message': 'Unexpected error occurred', 'error': str(e)}, status=500)

def asn_input(request):
    return render(request, 'kitting_config.html')


from django.contrib.auth.decorators import login_required
@login_required
def vc_list(request):
    vcs = EslPart.objects.all()
    vc_masters = VcMaster.objects.all()
    return render(request, 'model_matrix.html', {'vcs': vcs, 'vc_masters': vc_masters})

from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache

# @login_required
# @never_cache
# def Part_VC_Map_Data(request):
#     vcs = EslPart.objects.all()
#     vc_masters = VcMaster.objects.all()
#     return render(request,  "model_matrix/vc_part_map.html", {'vcs': vcs, 'vc_masters': vc_masters})

from django.shortcuts import redirect

@login_required
@never_cache
def Part_VC_Map_Data(request):
    if not request.user.is_authenticated:
        return redirect('login')
    vcs = EslPart.objects.all()
    vc_masters = VcMaster.objects.all()
    return render(request, "model_matrix/vc_part_map.html", {'vcs': vcs, 'vc_masters': vc_masters})
    
from django.contrib.auth.decorators import login_required
@login_required
def add_esl_part_template(request):
    return render(request, 'model_matrix/add_esl.html')   

def add_new_trolley_template(request):
    all_trolley_tags = trolley_data.objects.all()
    return render(request, 'model_matrix/add_new_trolley.html' , {'all_trolley_tags':all_trolley_tags})  
def vc_model_mapping_template(request):
    vc_masters = VcMaster.objects.all()
    return render(request, 'model_matrix/vc_model_mapping.html' ,{'vc_masters':vc_masters}) 

def fetch_existing_vc_entries(request):
    part_no = request.GET.get('part_no')
    if part_no:
        existing_vc_entries = VcDatabase.objects.filter(part_no=part_no).values_list('vc_no', flat=True)
        return JsonResponse(list(existing_vc_entries), safe=False)
    
    return JsonResponse([], safe=False)


def update_vc(request):
   
    if request.method == 'POST':
        vc_id = request.POST.get('vc_id')
        part_number = request.POST.get('part_number')
        part_desc = request.POST.get('part_desc')
        quantity = request.POST.get('quantity')
        selected_vc_ids = request.POST.getlist('vc_master_id[]')  # Use getlist to fetch multiple selected IDs

       

        try:
            existing_vc_entries = VcDatabase.objects.filter(part_no=part_number)
            existing_vc_ids = list(existing_vc_entries.values_list('vc_no', flat=True))

            selected_vc_numbers = list(VcMaster.objects.filter(id__in=selected_vc_ids).values_list('vcnumber', flat=True))

            # Create or update entries for checked items
            created_items = []
            for vc_id in selected_vc_ids:
                vc_master = VcMaster.objects.get(id=vc_id)
                

                if vc_master.vcnumber not in existing_vc_ids:
                    created_item = VcDatabase.objects.create(
                        vc_no=vc_master.vcnumber,
                        part_no=part_number,
                        part_desc=part_desc,
                        quantity=quantity
                    )
                    created_items.append(created_item)
                else:
                    vc_entry = VcDatabase.objects.get(vc_no=vc_master.vcnumber, part_no=part_number)
                    if vc_entry.quantity != int(quantity):
                        vc_entry.quantity = quantity
                        vc_entry.save()
                        # Update quantity in EslPart if the quantity has changed
                        try:
                            esl_part = EslPart.objects.get(part_no=part_number)
                            if esl_part.quantity != int(quantity):
                                esl_part.quantity = quantity
                                esl_part.save()
                        except EslPart.DoesNotExist:
                            EslPart.objects.create(partno=part_number, quantity=quantity)

            # Delete entries for unchecked items
            for vc_entry in existing_vc_entries:
                if vc_entry.vc_no not in selected_vc_numbers:
                    vc_entry.delete()

            # Show success message and redirect
            messages.success(request, 'VC entries updated successfully.')
            return redirect('Part_VC_Map_Data')   
            #return JsonResponse({'status': 'success', 'message': 'VC entries updated successfully.'})
            

        except Exception as e:
            # Handle exceptions and show error message
            messages.error(request, f'Error updating VC entries: {str(e)}')
            return redirect('Part_VC_Map_Data')
           # return JsonResponse({'status': 'error', 'message': f'Error updating VC entries: {str(e)}'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'})
def edit_vc(request, id):
    vc = get_object_or_404(VcDatabase, id=id)
    if request.method == "POST":
        form = VcDatabaseForm(request.POST, instance=vc)
        if form.is_valid():
            form.save()
            return redirect('vc_list')
    else:
        form = VcDatabaseForm(instance=vc)
    return render(request, 'edit_vc.html', {'form': form})

def delete_vc(request, part_no):
    try:
        esl_part = EslPart.objects.get(partno=part_no)
        if esl_part:
            esl_part.delete()
            return redirect('vc_list')
        return JsonResponse({'message': f'Part number {part_no} deleted from EslPart.'},status=204)
    except EslPart.DoesNotExist:
        return JsonResponse({'error': f'Part number {part_no} not found in EslPart.'}, status=404)
    
import logging
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import openpyxl
from .models import VcDatabase

logger = logging.getLogger(__name__)

@csrf_exempt
def add_esl_part(request):
    if request.method == 'POST':
        try:
            part_no = request.POST.get('part_no')
            tag_mac = request.POST.get('tagId')
            part_desc = request.POST.get('part_desc')
            quantity = request.POST.get('quantity')

            if not part_no or not part_desc or not quantity:
                return JsonResponse({'success': False, 'message': 'All fields are required.'})
            
            if  EslPart.objects.filter(partno=part_no).exists() or \
               EslPart.objects.filter(tagid=tag_mac).exists() or \
               EslPart.objects.filter(part_desc=part_desc).exists():
                return JsonResponse({'success': False, 'message': 'an entry already exist with the same parameter.'})

            EslPart.objects.create(
                   # Adjust as needed
                partno=part_no,
                part_desc=part_desc,
                tagid = tag_mac,
                quantity=quantity
            )

            return JsonResponse({'success': True, 'message': 'ESL Part added successfully.'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {e}'})

    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook

import pandas as pd
@csrf_exempt
def upload_esl_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        # Read the Excel file
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error reading Excel file: {str(e)}'})

        required_columns = ['part_no', 'tagId', 'part_desc', 'quantity']
        if not all(column in df.columns for column in required_columns):
            return JsonResponse({'success': False, 'message': 'Excel file must contain part_no, tagId, part_desc, and quantity columns'})

        skipped_entries = []
        for index, row in df.iterrows():
            part_no = row['part_no']
            tag_id = row['tagId']
            part_desc = row['part_desc']
            quantity = row['quantity']

            # Check if any of the fields (except quantity) match in the database
            if  EslPart.objects.filter(partno=part_no).exists() or \
               EslPart.objects.filter(tagid=tag_id).exists() or \
               EslPart.objects.filter(part_desc=part_desc).exists():
                skipped_entries.append(row.to_dict())
                messages.error(request , "item added succesfully but some item already exists with same data")
                
                continue

            # Add the ESL part to the database
            EslPart.objects.create(
                partno=part_no,
                tagid=tag_id,
                part_desc=part_desc,
                quantity=quantity
            )

        if skipped_entries:
            message = f'Some entries were not added because they already exist in the database: {skipped_entries}'
        else:
            message = 'All entries were added successfully.'
        messages.success(request , "items added to DB successfully.")
        return JsonResponse({'success': True, 'message': message})

    return JsonResponse({'success': False, 'message': 'Invalid request method or no file uploaded'})


@csrf_exempt
@require_http_methods(["POST"])
def upload_excel(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)

            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
                try:
                    vc_no, side, master_id, part_no, part_desc, quantity = row
                    logger.info(f"Processing row: {row}")  # Log the row being processed
                    existing_entries = VcDatabase.objects.filter(vc_no=vc_no, part_no=part_no)
                    if not existing_entries:
                        created= VcDatabase.objects.create(
                            part_no=part_no,
                            vc_no=vc_no,
                            side=side,
                            part_desc=part_desc,
                            quantity=quantity
                        )
                        logger.info(f"Created new entry for part_no: {part_no}, vc_no: {vc_no}")

                        if created:
                            logger.info(f"Created new entry for part_no: {part_no}")
                        else:
                            logger.info(f"Updated existing entry for part_no: {part_no}")

                except ValueError as e:
                    logger.error(f"ValueError processing row {row}: {e}")
                    continue  # Skip the row if there's a value error
                except Exception as e:
                    logger.error(f"Error processing row {row}: {e}")
                    continue  # Skip the row if there's any other error
            

            return JsonResponse({'success': True, 'message': 'Excel file processed successfully.'})

        except Exception as e:
            logger.error(f"Error processing the uploaded file: {e}")
            return JsonResponse({'success': False, 'message': f'There was an error processing the uploaded file: {e}'})

    return JsonResponse({'success': False, 'message': 'No file uploaded.'})


from django.shortcuts import render, redirect
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.forms import UserCreationForm
from .forms import SignUpForm, LoginForm

def user_register(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your account has been created successfully.')
            return redirect('/')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SignUpForm()
    return render(request, 'user_register.html', {'form': form})
def user_login(request):
    if request.method == 'POST':
        form = LoginForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('/')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = LoginForm()
    
    return render(request, 'user_login.html', {'form': form})

def user_logout(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    
    return redirect('login')

def download_sample_excel(request):
    filepath = os.path.join(os.path.dirname(__file__), 'sample.xlsx')
    with open(filepath, 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=sample.xlsx'
        return response

@require_http_methods(["DELETE"])
@csrf_exempt
def delete_vc_master(request, id):
    try:
        vc_master = VcMaster.objects.get(id=id)
        vc_num=vc_master.vcnumber
        # Delete related VC entries
        VcDatabase.objects.filter(vc_no=vc_num).delete()
        # Delete the VC master entry
        vc_master.delete()
        return JsonResponse({'success': True})
    except VcMaster.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'VC Master entry not found.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})
def add_vc_master(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        vcnumber = data.get('vcnumber')
        model = data.get('model')
        if vcnumber and model:
            new_vc_master = VcMaster(vcnumber=vcnumber, model=model)
            new_vc_master.save()
            return JsonResponse({'success': True, 'id': new_vc_master.id})
        else:
            return JsonResponse({'success': False}, status=400)

def add_trolley_esl(request):
    if request.method == 'POST':
        mac = request.POST.get('mac')
        trolley_code = request.POST.get('trolley_code')
        asn_num = request.POST.get('asn_num')
        color = request.POST.get('color')
        trolley_picking_status = request.POST.get('trolley_picking_status')

        try:
            trolley_data.objects.create(
                mac=mac,
                trolley_code=trolley_code,
                asn_num=asn_num,
                color=color,
                trolley_picking_status=trolley_picking_status
            )
            return JsonResponse({'success': True, 'message': 'Trolley ESL added successfully.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error adding Trolley ESL: {e}'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

from django.shortcuts import get_object_or_404
@csrf_exempt
def delete_trolley_esl(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            trolley_id = data.get('id')
            trolley = get_object_or_404(trolley_data, id=trolley_id)
            trolley.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request'})



def base_model_matrix(request):
    return render(request ,"model_matrix/base_model_matrix.html") 






@csrf_exempt
def callback_view(request):
    if request.method == "POST":
        try:
            raw_data = request.body.decode("utf-8")
            logger.info(f"Received callback data: {raw_data}")  # Log raw data

            # Log headers for debugging
            logger.info(f"Request headers: {request.headers}")

            data = json.loads(raw_data)
            logger.info(f"Parsed callback data: {data}")  # Log parsed data
            print(data)  # Print data to console for debugging

            return JsonResponse({"status": "success", "message": "Callback received"}, status=200)
        
        except json.JSONDecodeError as e:
            logger.error(f"JSON Decode Error: {e}")
            return JsonResponse({"status": "error", "message": "Invalid JSON format"}, status=400)

    logger.warning(f"Invalid request method: {request.method}")
    return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)




from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import requests

url2 = "http://192.168.1.100/wms/associate/lightTagsLed"
url3 = 'http://192.168.1.100/wms/associate/updateScreen'

@csrf_exempt  # Disable CSRF for testing purposes (enable CSRF protection in production)
def send_led_request(request):
    # Ensure only POST requests are processed
    
    # Define the data to send
    data_list = [
        {"mac":"92.94.88.81","mappingtype":682,"styleid":60,"Hello":"hhhello","ledrgb":"ff00","ledstate":"0","outtime":"0"},
           
    ]

    try:
        # Send the POST request
        response = requests.post(url3, json=data_list)
        response.raise_for_status()  # Raise an exception for HTTP errors
        print(f"Request Method: {request.method}")  # Should print "POST"

        return JsonResponse({
            "status": "success",
            "message": "Request sent successfully",
            "response": response.json()  # Include the response from the external API
        }, status=200)
    
    except requests.exceptions.RequestException as e:
        # Handle any errors that occur during the request
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)
# Call function to send request
#send_led_request()


from .models import AsnSchedule, WorkTable, vc_n_asn

def clear_tables(request):
    try:
        # Get the latest vc_n_asn entry
        latest_entry = vc_n_asn.objects.latest('id')
        
        if latest_entry.asnn == '999':
            # Delete all records from AsnSchedule
            AsnSchedule.objects.all().delete()
            
            # Delete all records from WorkTable
            WorkTable.objects.all().delete()
            
            return JsonResponse({
                'status': 'success',
                'message': 'All entries cleared successfully'
            })
        
        return JsonResponse({
            'status': 'info',
            'message': 'No clearing needed'
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

import pandas as pd
from django.http import HttpResponse
from .models import VcDatabase

import pandas as pd
from django.http import HttpResponse
from django.db.utils import OperationalError
from openpyxl.utils.exceptions import InvalidFileException

def export_vcdatabase_excel(request):
    try:
        # Attempt to retrieve data from the database
        queryset = VcDatabase.objects.all().values()
        
        # Check if the queryset is empty
        if not queryset:
            return HttpResponse("No data available to export.", status=404)

        # Create a DataFrame from the queryset
        df = pd.DataFrame(list(queryset))

    except OperationalError:
        # Handle database connection or query errors
        return HttpResponse("Database error: Could not retrieve data.", status=500)
    except Exception as e:
        # Catch any other unexpected errors during data retrieval
        return HttpResponse(f"An unexpected error occurred during data retrieval: {e}", status=500)

    try:
        # Prepare the HTTP response for the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=vc_database.xlsx'

        # Use a context manager to ensure the Excel writer is properly closed
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            # Write the DataFrame to the Excel file
            df.to_excel(writer, index=False)
        
        return response

    except InvalidFileException:
        # Handle errors related to the Excel file format
        return HttpResponse("Error generating the Excel file.", status=500)
    except Exception as e:
        # Handle any other unexpected errors during file generation
        return HttpResponse(f"An unexpected error occurred during file generation: {e}", status=500)
